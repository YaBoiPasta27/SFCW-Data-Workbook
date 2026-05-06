"""
SF Climate Week Event Scraper
Scrapes events from:
  - Climatebase API (2026): https://sfcw.climate-week.org/
  - Luma calendar proxy (2025): https://www.sfclimateweek.org/events/all
and exports to formatted Excel spreadsheets.

Requirements:
    pip install requests openpyxl
"""

import base64
import json
import time
from datetime import datetime, timezone
from typing import Dict, List, Optional

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Config ────────────────────────────────────────────────────────────────────

# 2026 – Climatebase API
SLUG = "sfcw-26"
BASE_API = "https://api.climatebase.org/api/v1/climate_week/public/cities"

# 2023/2024/2025 – Luma calendars (via sfclimateweek.org proxy)
LUMA_PROXY = "https://www.sfclimateweek.org/api/luma-events"
LUMA_CALENDAR_2025 = "cal-NvjGvvobq9xsnqs"
LUMA_CALENDAR_2024 = "cal-KJKVd0YCMFZ6ddZ"
LUMA_CALENDAR_2023 = "cal-FYQxwFCJ7Hczj0Z"

OUTPUT_FILE_ALL  = "sfcw_events_all.xlsx"
OUTPUT_FILE_2026 = "sfcw_events_2026.xlsx"
OUTPUT_FILE_2025 = "sfcw_events_2025.xlsx"
OUTPUT_FILE_2024 = "sfcw_events_2024.xlsx"
OUTPUT_FILE_2023 = "sfcw_events_2023.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json",
    "Referer": "https://sfcw.climate-week.org/",
    "Origin": "https://sfcw.climate-week.org",
}

COLUMNS = [
    "Title",
    "Date",
    "Location",
    "City",
    "Event Type",
    "Description",
    "Host/Organizer",
    "Cost",
    "Speaker(s)",
    "Virtual/In-Person",
]


# ── Climatebase API calls (2026) ───────────────────────────────────────────────

def fetch_distinct_dates() -> List[str]:
    """Fetch the list of dates that have events."""
    url = f"{BASE_API}/events/distinct-dates?slug={SLUG}"
    print(f"  Fetching distinct dates from: {url}")
    r = requests.get(url, headers=HEADERS, timeout=15)
    r.raise_for_status()
    data = r.json()
    print(f"  Raw distinct-dates response keys: {list(data.keys())}")
    dates = data.get("distinct_dates", [])
    print(f"  Found {len(dates)} distinct dates: {dates}")
    return dates


def fetch_events_for_date(date_str: str) -> List[dict]:
    """Fetch all events for a single date, paginating until exhausted."""
    all_events: List[dict] = []
    limit = 100  # API hard cap is 100 regardless of the limit param
    offset = 0
    total = None
    while True:
        url = (
            f"{BASE_API}/events"
            f"?slug={SLUG}&status=approved&limit={limit}&offset={offset}"
            f"&window_start={date_str}&window_end={date_str}"
        )
        print(f"\n  Fetching events for {date_str} (offset={offset}): {url}")
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        data = r.json()
        if offset == 0:
            print(f"  Response keys: {list(data.keys())}")
            total = data.get("total_events_for_query")
            print(f"  total_events_for_query: {total}")
        events = data.get("events", [])
        all_events.extend(events)
        print(f"  Events returned this page: {len(events)}")
        offset += len(events)
        if not events or (total is not None and offset >= total):
            break
        time.sleep(0.3)
    if all_events:
        print(f"  Sample event keys: {list(all_events[0].keys())}")
    return all_events


def fetch_spotlight_events() -> List[dict]:
    """Fetch flagship/spotlight events (shown at top of page)."""
    url = f"{BASE_API}/events?slug={SLUG}&status=approved&limit=100&offset=0&only_spotlight=true"
    print(f"\n  Fetching spotlight events: {url}")
    r = requests.get(url, headers=HEADERS, timeout=15)
    r.raise_for_status()
    data = r.json()
    events = data.get("events", [])
    print(f"  Spotlight events returned: {len(events)}")
    return events


# ── Luma API calls ────────────────────────────────────────────────────────────

def _make_luma_cursor(iso_datetime: str) -> str:
    """Encode a Luma pagination cursor that starts just before the given UTC datetime."""
    payload = {"sv": iso_datetime, "fb": ""}
    return base64.b64encode(json.dumps(payload, separators=(",", ":")).encode()).decode().rstrip("=")


def fetch_luma_events(
    calendar_id: str,
    start_date: str,
    end_date: str,
) -> List[dict]:
    """
    Fetch all Luma calendar entries between start_date and end_date (inclusive, SF local time).
    Uses a crafted cursor to jump directly to the right date window.
    """
    start_cursor = _make_luma_cursor(f"{end_date}T23:59:59.000Z")
    cutoff_utc = f"{start_date}T00:00:00.000Z"

    all_entries: List[dict] = []
    cursor: Optional[str] = start_cursor
    page = 0

    print(f"\n  Fetching Luma events ({start_date} – {end_date}) from {calendar_id}...")

    while True:
        params = {
            "calendar_api_id": calendar_id,
            "period": "past",
            "pagination_limit": 100,
        }
        if cursor:
            params["pagination_cursor"] = cursor

        r = requests.get(LUMA_PROXY, params=params, timeout=20)
        r.raise_for_status()
        data = r.json()

        entries = data.get("entries", [])
        has_more = data.get("has_more", False)
        cursor = data.get("next_cursor")
        page += 1

        # Filter to the requested date window and check stopping condition
        page_in_range: List[dict] = []
        passed_cutoff = False
        for entry in entries:
            start_at = entry.get("event", {}).get("start_at") or entry.get("start_at") or ""
            if start_at < cutoff_utc:
                passed_cutoff = True
                break
            page_in_range.append(entry)

        all_entries.extend(page_in_range)
        print(f"  Page {page}: {len(entries)} returned, {len(page_in_range)} in range "
              f"(total so far: {len(all_entries)}, has_more={has_more})")

        if passed_cutoff or not has_more or not cursor:
            break
        time.sleep(0.25)

    print(f"  Total Luma entries in date range: {len(all_entries)}")
    return all_entries


# ── Parsing ───────────────────────────────────────────────────────────────────

def normalize_event_payload(item: dict) -> dict:
    """Flatten nested event structures from both Climatebase and Luma formats."""
    if not (isinstance(item, dict) and isinstance(item.get("event"), dict)):
        return item

    nested = dict(item["event"])

    # Common top-level fields
    nested["spotlight"] = item.get("spotlight")
    nested["tag_categories"] = item.get("tag_categories")

    # Luma: hosts array lives at entry level
    if "hosts" in item:
        nested.setdefault("hosts", item["hosts"])

    # Luma: ticket_info
    ti = item.get("ticket_info")
    if isinstance(ti, dict):
        nested.setdefault("is_free", ti.get("is_free"))
        price = ti.get("price")
        if isinstance(price, dict) and price.get("cents") is not None:
            # Convert {"cents": 2500, "currency": "usd"} → "$25.00"
            cents = price["cents"]
            symbol = {"usd": "$", "eur": "€", "gbp": "£"}.get(
                (price.get("currency") or "usd").lower(), "$"
            )
            nested.setdefault("cost", f"{symbol}{cents / 100:.2f}")
        elif price is not None:
            nested.setdefault("cost", str(price))

    # Luma: tags → event type (more useful than the generic "independent" event_type)
    # Strip format/attendance tags; those belong in the Virtual/In-Person column.
    _FORMAT_TAGS = {"in-person event", "online event", "virtual event", "hybrid event"}
    tags = item.get("tags") or nested.get("tags") or []
    tag_names: List[str] = []
    if tags and isinstance(tags, list):
        nested.setdefault("tags", tags)
        tag_names = [
            t.get("name") or t.get("label") or ""
            for t in tags if isinstance(t, dict)
            if (t.get("name") or "").lower() not in _FORMAT_TAGS
        ]
    tag_str = ", ".join(t for t in tag_names if t)
    # Always override Luma's generic "independent" placeholder (even when tags is empty)
    if nested.get("event_type") in (None, "", "independent"):
        nested["event_type"] = tag_str
    else:
        nested.setdefault("event_type", tag_str)

    # Luma: flatten geo_address_info
    geo = nested.get("geo_address_info")
    if isinstance(geo, dict):
        nested.setdefault("full_address", geo.get("full_address") or geo.get("address"))
        nested.setdefault("city", geo.get("city_name") or geo.get("city"))

    # Luma: location_type → virtual/in-person flags
    lt = nested.get("location_type") or ""
    virtual_types = {"online", "zoom", "virtual", "livestream"}
    if lt in virtual_types:
        nested.setdefault("is_online", True)
    elif lt == "hybrid":
        nested.setdefault("is_virtual", True)
        nested.setdefault("is_in_person", True)
    elif lt == "offline" or (not lt and nested.get("geo_address_info")):
        nested.setdefault("is_online", False)

    return nested


def parse_event(item: dict) -> dict:
    """Map a raw API event dict to our spreadsheet columns."""
    item = normalize_event_payload(item)
    print(f"    Parsing: {(item.get('title') or item.get('name') or item.get('event_name') or '(no title)')!r}")

    # Date/time
    raw_start = (
        item.get("start_date")
        or item.get("start_time")
        or item.get("date")
        or item.get("datetime")
        or item.get("start_at")
        or item.get("startAt")
        or ""
    )
    raw_end = (
        item.get("end_date")
        or item.get("end_time")
        or item.get("end_at")
        or item.get("endAt")
        or ""
    )
    date_str = ""
    if raw_start:
        try:
            dt = datetime.fromisoformat(str(raw_start).replace("Z", "+00:00"))
            date_str = dt.strftime("%b %d, %Y %I:%M %p")
            if raw_end:
                dt_end = datetime.fromisoformat(str(raw_end).replace("Z", "+00:00"))
                date_str += " - " + dt_end.strftime("%I:%M %p")
        except Exception:
            date_str = str(raw_start)

    # Virtual / In-Person
    fmt = (item.get("format") or item.get("mode") or item.get("type") or "").lower()
    is_virtual = (
        item.get("is_virtual")
        or item.get("online")
        or item.get("is_online")
        or "virtual" in fmt
        or "online" in fmt
    )
    is_in_person = (
        item.get("is_in_person")
        or (item.get("is_online") is False and bool(item.get("full_address") or item.get("address")))
        or "in-person" in fmt
        or "person" in fmt
    )
    if is_virtual and is_in_person:
        format_str = "Hybrid"
    elif is_virtual:
        format_str = "Virtual"
    elif is_in_person:
        format_str = "In-Person"
    else:
        format_str = item.get("format") or item.get("mode") or ""

    # Location
    location_parts = [
        item.get("venue_name") or item.get("venue") or "",
        item.get("full_address") or item.get("address") or item.get("location") or "",
    ]
    location = ", ".join(p for p in location_parts if p) or item.get("location") or ""

    # Speakers / hosts
    speakers_raw = item.get("speakers") or item.get("hosts") or item.get("host_names") or []
    if isinstance(speakers_raw, list):
        speakers = ", ".join(
            (s.get("name") or s.get("full_name") or str(s)) if isinstance(s, dict) else str(s)
            for s in speakers_raw if s
        )
    else:
        speakers = str(speakers_raw)

    # Cost
    cost_raw = item.get("cost") or item.get("price") or item.get("ticket_price") or ""
    cost_str = str(cost_raw).strip()
    if item.get("is_free") or item.get("is_paid") is False or cost_str in ("0", "0.0", "0.00", ""):
        cost = "Free" if (item.get("is_free") or item.get("is_paid") is False or cost_str in ("0", "0.0", "0.00")) else ""
    else:
        cost = cost_str

    return {
        "Title": (
            item.get("title")
            or item.get("name")
            or item.get("event_name")
            or item.get("event_title")
            or ""
        ),
        "Date": date_str,
        "Location": location,
        "City": item.get("city") or item.get("city_name") or "San Francisco",
        "Event Type": (
            item.get("event_type")
            or item.get("type")
            or item.get("category")
            or ""
        ),
        "Description": item.get("description") or item.get("summary") or item.get("details") or "",
        "Host/Organizer": (
            item.get("organizer")
            or item.get("host")
            or item.get("organization")
            or item.get("organization_name")
            or speakers
            or ""
        ),
        "Cost": cost,
        "Speaker(s)": speakers,
        "Virtual/In-Person": format_str,
    }


# ── Scrapers ──────────────────────────────────────────────────────────────────

def run_scraper_2026() -> List[Dict]:
    print("\n" + "=" * 60)
    print("  SF Climate Week 2026  (Climatebase API)")
    print("=" * 60)
    print(f"  Slug: {SLUG}")
    print(f"  API:  {BASE_API}")
    print()

    all_raw: List[dict] = []

    print("Step 1: Fetching distinct event dates...")
    try:
        dates = fetch_distinct_dates()
    except Exception as e:
        print(f"  ERROR fetching dates: {e}")
        dates = []

    if dates:
        print(f"\nStep 2: Fetching events for {len(dates)} date(s)...")
        for date_str in dates:
            try:
                events = fetch_events_for_date(date_str)
                all_raw.extend(events)
            except Exception as e:
                print(f"  ERROR fetching {date_str}: {e}")
            time.sleep(0.3)
    else:
        print("\nStep 2: No dates found — trying undated fetch...")
        url = f"{BASE_API}/events?slug={SLUG}&status=approved&limit=500&offset=0"
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            data = r.json()
            print(f"  Response keys: {list(data.keys())}")
            all_raw = data.get("events", [])
            print(f"  Events returned: {len(all_raw)}")
        except Exception as e:
            print(f"  ERROR: {e}")

    print("\nStep 3: Fetching spotlight/flagship events...")
    try:
        spotlight = fetch_spotlight_events()
        all_raw.extend(spotlight)
    except Exception as e:
        print(f"  ERROR fetching spotlight events: {e}")

    print(f"\n  Total raw events collected (before dedup): {len(all_raw)}")

    print("\nStep 4: Parsing events...")
    seen = set()
    events: List[Dict] = []
    for item in all_raw:
        parsed = parse_event(item)
        key = (parsed["Title"].strip().lower(), parsed["Date"])
        if parsed["Title"].strip() and key not in seen:
            seen.add(key)
            events.append(parsed)

    print(f"\n  Total unique events: {len(events)}")
    return events


def _run_luma_scraper(year: str, calendar_id: str, start_date: str, end_date: str) -> List[Dict]:
    print("\n" + "=" * 60)
    print(f"  SF Climate Week {year}  (Luma calendar)")
    print("=" * 60)
    print(f"  Calendar: {calendar_id}")
    print(f"  Date range: {start_date} – {end_date}")
    print()

    print("Step 1: Fetching Luma events...")
    try:
        raw_entries = fetch_luma_events(calendar_id, start_date, end_date)
    except Exception as e:
        print(f"  ERROR: {e}")
        raw_entries = []

    print(f"\nStep 2: Parsing {len(raw_entries)} entries...")
    seen = set()
    events: List[Dict] = []
    for entry in raw_entries:
        parsed = parse_event(entry)
        key = (parsed["Title"].strip().lower(), parsed["Date"])
        if parsed["Title"].strip() and key not in seen:
            seen.add(key)
            events.append(parsed)

    print(f"\n  Total unique events: {len(events)}")
    return events


def run_scraper_2025(
    start_date: str = "2025-04-18",
    end_date: str = "2025-04-27",
) -> List[Dict]:
    return _run_luma_scraper("2025", LUMA_CALENDAR_2025, start_date, end_date)


def run_scraper_2024(
    start_date: str = "2024-04-21",
    end_date: str = "2024-04-27",
) -> List[Dict]:
    return _run_luma_scraper("2024", LUMA_CALENDAR_2024, start_date, end_date)


def run_scraper_2023(
    start_date: str = "2023-04-17",
    end_date: str = "2023-04-23",
) -> List[Dict]:
    return _run_luma_scraper("2023", LUMA_CALENDAR_2023, start_date, end_date)


# ── Excel export ──────────────────────────────────────────────────────────────

HEADER_COLOR  = "1A5276"
ALT_ROW_COLOR = "EBF5FB"
WHITE         = "FFFFFF"

COLUMNS_WITH_YEAR = ["Year"] + COLUMNS

COL_WIDTHS = {
    "Year": 7,
    "Title": 42, "Date": 24, "Location": 36, "City": 18,
    "Event Type": 20, "Description": 58, "Host/Organizer": 32,
    "Cost": 12, "Speaker(s)": 36, "Virtual/In-Person": 17,
}


def _write_sheet(
    ws,
    events: List[Dict],
    columns: List[str],
    year_col: bool = False,
):
    """Write a formatted events table to an existing worksheet."""
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(columns)
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = Font(name="Arial", bold=True, color=WHITE, size=11)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR, end_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 32

    alt_fill   = PatternFill("solid", start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR)
    white_fill = PatternFill("solid", start_color=WHITE, end_color=WHITE)

    for row_idx, event in enumerate(events, 2):
        row_vals = [event.get(col, "") for col in columns]
        ws.append(row_vals)
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill      = fill
            cell.border    = border

    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 20)
    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_excel(events: List[Dict], output_path: str, year_label: str = ""):
    """Write a single-year spreadsheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SFCW Events"
    _write_sheet(ws, events, COLUMNS)

    ws2 = wb.create_sheet("Summary")
    title = f"SF Climate Week {year_label} - Scrape Summary" if year_label else "SF Climate Week - Scrape Summary"
    ws2["A1"] = title
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)
    n = len(events)
    summary_rows = [
        ("Generated",        datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total Events",     n),
        ("Virtual Events",   f"=COUNTIF('SFCW Events'!J2:J{n+1},\"Virtual\")"),
        ("In-Person Events", f"=COUNTIF('SFCW Events'!J2:J{n+1},\"In-Person\")"),
        ("Hybrid Events",    f"=COUNTIF('SFCW Events'!J2:J{n+1},\"Hybrid\")"),
        ("Free Events",      f"=COUNTIF('SFCW Events'!H2:H{n+1},\"Free\")"),
    ]
    for i, (label, value) in enumerate(summary_rows, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=i, column=2, value=value).font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 42

    wb.save(output_path)
    print(f"\n  Saved: {output_path}")


def build_combined_excel(
    events_by_year: Dict[str, List[Dict]],
    output_path: str,
):
    """Write all years into one workbook: an 'All Events' sheet + per-year sheets + summary."""
    wb = openpyxl.Workbook()

    # All Events sheet (sorted oldest → newest, with Year column)
    ws_all = wb.active
    ws_all.title = "All Events"
    all_events = []
    for year, events in sorted(events_by_year.items()):
        for ev in events:
            all_events.append({**ev, "Year": year})
    all_events.sort(key=lambda e: e.get("Date") or "")
    _write_sheet(ws_all, all_events, COLUMNS_WITH_YEAR)

    # Per-year sheets
    for year, events in sorted(events_by_year.items()):
        ws_yr = wb.create_sheet(title=f"SFCW {year}")
        _write_sheet(ws_yr, events, COLUMNS)

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "SF Climate Week – All Years Summary"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=14)
    headers = ["Year", "Events", "In-Person", "Virtual", "Hybrid", "Free"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=3, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", start_color=HEADER_COLOR, end_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")

    row = 4
    total_events = 0
    for year in sorted(events_by_year.keys()):
        sheet_name = f"SFCW {year}"
        n = len(events_by_year[year])
        total_events += n
        j_range = f"'{sheet_name}'!J2:J{n+1}"
        h_range = f"'{sheet_name}'!H2:H{n+1}"
        ws_sum.cell(row=row, column=1, value=year).font       = Font(name="Arial", bold=True, size=10)
        ws_sum.cell(row=row, column=2, value=n).font          = Font(name="Arial", size=10)
        ws_sum.cell(row=row, column=3, value=f'=COUNTIF({j_range},"In-Person")').font  = Font(name="Arial", size=10)
        ws_sum.cell(row=row, column=4, value=f'=COUNTIF({j_range},"Virtual")').font    = Font(name="Arial", size=10)
        ws_sum.cell(row=row, column=5, value=f'=COUNTIF({j_range},"Hybrid")').font     = Font(name="Arial", size=10)
        ws_sum.cell(row=row, column=6, value=f'=COUNTIF({h_range},"Free")').font       = Font(name="Arial", size=10)
        row += 1

    # Totals row
    ws_sum.cell(row=row, column=1, value="TOTAL").font  = Font(name="Arial", bold=True, size=10)
    ws_sum.cell(row=row, column=2, value=total_events).font = Font(name="Arial", bold=True, size=10)
    for col in range(3, 7):
        ws_sum.cell(row=row, column=col,
            value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{row-1})"
        ).font = Font(name="Arial", bold=True, size=10)

    ws_sum.cell(row=row + 2, column=1, value="Generated").font = Font(name="Arial", bold=True, size=10)
    ws_sum.cell(row=row + 2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M")).font = Font(name="Arial", size=10)
    for col_idx, width in enumerate([10, 12, 12, 12, 10, 10], 1):
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    print(f"\n  Saved: {output_path}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    import sys
    mode = sys.argv[1] if len(sys.argv) > 1 else "all"

    SCRAPERS = {
        "2026": (run_scraper_2026, OUTPUT_FILE_2026),
        "2025": (run_scraper_2025, OUTPUT_FILE_2025),
        "2024": (run_scraper_2024, OUTPUT_FILE_2024),
        "2023": (run_scraper_2023, OUTPUT_FILE_2023),
    }

    years_to_run = sorted(SCRAPERS.keys()) if mode == "all" else [mode]
    events_by_year: Dict[str, List[Dict]] = {}

    for year in years_to_run:
        if year not in SCRAPERS:
            print(f"  Unknown year: {year}. Valid: {', '.join(sorted(SCRAPERS))}")
            continue
        scraper_fn, out_file = SCRAPERS[year]
        events = scraper_fn()
        if events:
            build_excel(events, out_file, year_label=year)
            print(f"  {len(events)} events → {out_file}")
            events_by_year[year] = events
        else:
            print(f"\n  No {year} events found.")

    if len(events_by_year) > 1:
        print("\nBuilding combined spreadsheet...")
        build_combined_excel(events_by_year, OUTPUT_FILE_ALL)
        total = sum(len(v) for v in events_by_year.values())
        print(f"  {total} total events across {len(events_by_year)} years → {OUTPUT_FILE_ALL}")

    print("\n" + "=" * 60)
    print("  All done!")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
